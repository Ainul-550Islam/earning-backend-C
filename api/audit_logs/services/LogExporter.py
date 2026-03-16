"""
Service for exporting audit logs in various formats
"""

import json
import csv
import io
import zipfile
import gzip
import tempfile
import os
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, BinaryIO, Union
from django.db.models import QuerySet
from django.utils import timezone
from django.http import HttpResponse
from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, Image, PageTemplate, Frame,
    BaseDocTemplate, NextPageTemplate
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from ..models import AuditLog, AuditLogArchive
from ..serializers import AuditLogSerializer, AuditLogDetailSerializer

logger = logging.getLogger(__name__)


class LogExporter:
    """Export audit logs in various formats"""
    
    DEFAULT_EXPORT_FIELDS = [
        'id', 'timestamp', 'action', 'level', 'user__email',
        'user_ip', 'message', 'status_code', 'success',
        'response_time_ms', 'resource_type', 'resource_id'
    ]
    
    FIELD_MAPPINGS = {
        'id': 'ID',
        'timestamp': 'Timestamp',
        'action': 'Action',
        'level': 'Level',
        'user__email': 'User Email',
        'user__username': 'Username',
        'user_ip': 'IP Address',
        'message': 'Message',
        'status_code': 'Status Code',
        'success': 'Success',
        'error_message': 'Error Message',
        'response_time_ms': 'Response Time (ms)',
        'resource_type': 'Resource Type',
        'resource_id': 'Resource ID',
        'country': 'Country',
        'city': 'City',
        'request_method': 'HTTP Method',
        'request_path': 'Request Path',
        'correlation_id': 'Correlation ID',
        'created_at': 'Created At',
    }
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
        
    def export(self, queryset: QuerySet, format: str = 'json', 
               fields: List[str] = None, include_related: bool = False,
               compression: str = 'none') -> Union[dict, io.BytesIO, HttpResponse]:
        """
        Export audit logs in specified format
        
        Args:
            queryset: QuerySet of audit logs
            format: Export format ('json', 'csv', 'excel', 'pdf')
            fields: List of fields to include
            include_related: Include related objects
            compression: Compression type ('none', 'gzip', 'zip')
        
        Returns:
            Export data in requested format
        """
        if not fields:
            fields = self.DEFAULT_EXPORT_FIELDS
        
        # Get data from queryset
        data = self._prepare_data(queryset, fields, include_related)
        
        # Export in requested format
        if format == 'json':
            result = self._export_json(data)
        elif format == 'csv':
            result = self._export_csv(data, fields)
        elif format == 'excel':
            result = self._export_excel(data, fields)
        elif format == 'pdf':
            result = self._export_pdf(data, fields)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        # Apply compression if requested
        if compression != 'none':
            result = self._compress_data(result, compression, format)
        
        return result
    
    def _prepare_data(self, queryset: QuerySet, fields: List[str], 
                     include_related: bool = False) -> List[Dict]:
        """Prepare data for export"""
        logs = list(queryset)
        
        if include_related:
            serializer_class = AuditLogDetailSerializer
        else:
            serializer_class = AuditLogSerializer
        
        serializer = serializer_class(logs, many=True)
        data = serializer.data
        
        # Filter fields if specified
        if fields != ['__all__']:
            filtered_data = []
            for item in data:
                filtered_item = {}
                for field in fields:
                    # Handle nested fields
                    if '__' in field:
                        value = item
                        for subfield in field.split('__'):
                            if isinstance(value, dict):
                                value = value.get(subfield, '')
                            else:
                                value = getattr(value, subfield, '')
                                if callable(value):
                                    value = value()
                        filtered_item[field] = value
                    else:
                        filtered_item[field] = item.get(field, '')
                filtered_data.append(filtered_item)
            data = filtered_data
        
        return data
    
    def _export_json(self, data: List[Dict]) -> dict:
        """Export as JSON"""
        return {
            'metadata': {
                'export_date': timezone.now().isoformat(),
                'total_records': len(data),
                'format': 'json',
            },
            'logs': data
        }
    
    def _export_csv(self, data: List[Dict], fields: List[str]) -> io.StringIO:
        """Export as CSV"""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields)
        
        # Write header with friendly names
        header = {field: self.FIELD_MAPPINGS.get(field, field.replace('_', ' ').title()) 
                  for field in fields}
        writer.writerow(header)
        
        # Write data rows
        for row in data:
            # Convert dict values to strings
            csv_row = {}
            for field in fields:
                value = row.get(field, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                elif value is None:
                    value = ''
                csv_row[field] = str(value)
            writer.writerow(csv_row)
        
        output.seek(0)
        return output
    
    def _export_excel(self, data: List[Dict], fields: List[str]) -> io.BytesIO:
        """Export as Excel (XLSX)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = self.FIELD_MAPPINGS.get(field, field.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, log in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                value = log.get(field, '')
                
                # Format special fields
                if field == 'timestamp' and value:
                    try:
                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        value = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except (ValueError, AttributeError):
                        pass
                
                # Handle nested values
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)[:32767]  # Excel cell limit
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value if value is not None else ''
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        self._create_summary_sheet(wb, data, fields)
        
        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict], fields: List[str]):
        """Create summary sheet in Excel workbook"""
        ws_summary = wb.create_sheet(title="Summary")
        
        # Summary statistics
        total_records = len(data)
        success_count = sum(1 for log in data if log.get('success') is True)
        error_count = sum(1 for log in data if log.get('level') == 'ERROR')
        
        # By level distribution
        level_counts = {}
        for log in data:
            level = log.get('level', 'UNKNOWN')
            level_counts[level] = level_counts.get(level, 0) + 1
        
        # By action distribution
        action_counts = {}
        for log in data:
            action = log.get('action', 'UNKNOWN')
            action_counts[action] = action_counts.get(action, 0) + 1
        
        # Write summary
        ws_summary['A1'] = "Audit Logs Export Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        ws_summary['A3'] = "Export Date:"
        ws_summary['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws_summary['A4'] = "Total Records:"
        ws_summary['B4'] = total_records
        
        ws_summary['A5'] = "Success Rate:"
        ws_summary['B5'] = f"{(success_count / total_records * 100):.1f}%" if total_records > 0 else "N/A"
        
        ws_summary['A6'] = "Error Count:"
        ws_summary['B6'] = error_count
        
        # Level distribution
        ws_summary['A8'] = "Level Distribution"
        ws_summary['A8'].font = Font(bold=True)
        
        row = 9
        for level, count in sorted(level_counts.items()):
            ws_summary[f'A{row}'] = level
            ws_summary[f'B{row}'] = count
            row += 1
        
        # Action distribution
        ws_summary['D8'] = "Action Distribution"
        ws_summary['D8'].font = Font(bold=True)
        
        row = 9
        for action, count in sorted(action_counts.items(), key=lambda x: x[1], reverse=True)[:10]:  # Top 10
            ws_summary[f'D{row}'] = action
            ws_summary[f'E{row}'] = count
            row += 1
    
    def _export_pdf(self, data: List[Dict], fields: List[str]) -> io.BytesIO:
        """Export as PDF"""
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Define styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.white,
            alignment=TA_CENTER
        )
        
        # Build story
        story = []
        
        # Title
        title = Paragraph("Audit Logs Export", title_style)
        story.append(title)
        
        # Metadata
        metadata = [
            f"Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Total Records: {len(data)}",
            f"Generated By: PDF Export System"
        ]
        
        for meta in metadata:
            story.append(Paragraph(meta, styles['Normal']))
            story.append(Spacer(1, 5))
        
        story.append(Spacer(1, 20))
        
        # Prepare table data
        table_data = []
        
        # Header row
        header_row = [self.FIELD_MAPPINGS.get(field, field.replace('_', ' ').title()) 
                     for field in fields]
        table_data.append(header_row)
        
        # Data rows (limit to 1000 rows for PDF performance)
        max_rows = 1000
        for i, log in enumerate(data[:max_rows]):
            row = []
            for field in fields:
                value = log.get(field, '')
                
                # Format values
                if field == 'timestamp' and value:
                    try:
                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        value = dt.strftime('%Y-%m-%d\n%H:%M:%S')
                    except (ValueError, AttributeError):
                        pass
                
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)[:100] + '...'
                
                if value is None:
                    value = ''
                
                row.append(str(value))
            
            # Alternate row colors
            if i % 2 == 0:
                row_style = ('BACKGROUND', (0, i+1), (-1, i+1), colors.lightgrey)
            else:
                row_style = None
            
            table_data.append(row)
        
        # Create table
        table = Table(table_data, repeatRows=1)
        
        # Style table
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
            
            ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ])
        
        # Add alternate row colors
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
        
        table.setStyle(table_style)
        
        # Add table to story
        story.append(table)
        
        # Add page break if there are more rows
        if len(data) > max_rows:
            story.append(PageBreak())
            story.append(Paragraph(f"Continued... (showing first {max_rows} of {len(data)} records)", 
                                  styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        return buffer
    
    def _compress_data(self, data: Union[dict, io.IOBase], compression: str, 
                      original_format: str) -> io.BytesIO:
        """Compress data using specified compression method"""
        
        if compression == 'gzip':
            return self._compress_gzip(data, original_format)
        elif compression == 'zip':
            return self._compress_zip(data, original_format)
        else:
            raise ValueError(f"Unsupported compression: {compression}")
    
    def _compress_gzip(self, data: Union[dict, io.IOBase], original_format: str) -> io.BytesIO:
        """Compress data using GZIP"""
        
        buffer = io.BytesIO()
        
        with gzip.GzipFile(fileobj=buffer, mode='wb') as f:
            if isinstance(data, dict):  # JSON
                json_bytes = json.dumps(data, indent=2, default=str).encode('utf-8')
                f.write(json_bytes)
            elif hasattr(data, 'getvalue'):  # BytesIO or StringIO
                if isinstance(data, io.StringIO):
                    f.write(data.getvalue().encode('utf-8'))
                else:
                    f.write(data.getvalue())
            elif isinstance(data, str):
                f.write(data.encode('utf-8'))
        
        buffer.seek(0)
        return buffer
    
    def _compress_zip(self, data: Union[dict, io.IOBase], original_format: str) -> io.BytesIO:
        """Compress data using ZIP"""
        
        buffer = io.BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            filename = f"audit_logs_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
            
            if original_format == 'json':
                filename += '.json'
                if isinstance(data, dict):
                    content = json.dumps(data, indent=2, default=str)
                else:
                    content = json.dumps(data, indent=2, default=str)
                zip_file.writestr(filename, content)
            
            elif original_format == 'csv':
                filename += '.csv'
                if hasattr(data, 'getvalue'):
                    content = data.getvalue()
                    if isinstance(content, str):
                        content = content.encode('utf-8')
                else:
                    content = str(data).encode('utf-8')
                zip_file.writestr(filename, content)
            
            elif original_format == 'excel':
                filename += '.xlsx'
                if hasattr(data, 'getvalue'):
                    content = data.getvalue()
                else:
                    content = bytes(data)
                zip_file.writestr(filename, content)
            
            elif original_format == 'pdf':
                filename += '.pdf'
                if hasattr(data, 'getvalue'):
                    content = data.getvalue()
                else:
                    content = bytes(data)
                zip_file.writestr(filename, content)
            
            # Add README file
            readme_content = f"""Audit Logs Export
=================

Export Details:
- Format: {original_format.upper()}
- Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S %Z')}
- Compression: ZIP
- Contents: {filename}

This file contains audit logs exported from the system.
"""
            zip_file.writestr('README.txt', readme_content)
        
        buffer.seek(0)
        return buffer
    
    def create_archive(self, start_date=None, end_date=None, compression='zip'):
        """
        Create archive of old logs and store in database
        
        Args:
            start_date: Start date for archive
            end_date: End date for archive
            compression: Compression type
        
        Returns:
            AuditLogArchive instance
        """
        # Set default dates
        if not end_date:
            end_date = timezone.now() - timedelta(days=30)  # Default: logs older than 30 days
        
        if not start_date:
            # Archive all logs before end_date
            queryset = AuditLog.objects.filter(timestamp__lt=end_date, archived=False)
        else:
            queryset = AuditLog.objects.filter(
                timestamp__range=(start_date, end_date),
                archived=False
            )
        
        total_logs = queryset.count()
        
        if total_logs == 0:
            raise ValueError("No logs to archive")
        
        # Export logs
        export_data = self.export(
            queryset=queryset,
            format='json',
            fields=['__all__'],
            compression=compression
        )
        
        # Calculate sizes
        if hasattr(export_data, 'getvalue'):
            compressed_size = len(export_data.getvalue()) / (1024 * 1024)  # MB
        else:
            compressed_size = len(export_data) / (1024 * 1024)
        
        # Estimate original size (rough estimate)
        original_size = total_logs * 2 / 1024  # ~2KB per log
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        if hasattr(export_data, 'getvalue'):
            temp_file.write(export_data.getvalue())
        else:
            temp_file.write(export_data)
        temp_file.close()
        
        # Create archive record
        archive = AuditLogArchive.objects.create(
            start_date=start_date or queryset.earliest('timestamp').timestamp,
            end_date=end_date,
            total_logs=total_logs,
            compressed_size_mb=compressed_size,
            original_size_mb=original_size,
            compression_ratio=compressed_size / original_size if original_size > 0 else 0,
            storage_path=temp_file.name
        )
        
        # Mark logs as archived
        queryset.update(archived=True)
        
        logger.info(f"Created archive {archive.id} with {total_logs} logs")
        
        return archive
    
    def export_for_api(self, queryset: QuerySet, format: str = 'json', 
                      fields: List[str] = None) -> HttpResponse:
        """
        Export data as HTTP response for API
        
        Args:
            queryset: QuerySet of logs
            format: Export format
            fields: Fields to include
        
        Returns:
            HTTP response
        """
        data = self.export(queryset, format, fields)
        
        filename = f"audit_logs_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format == 'json':
            response = HttpResponse(
                json.dumps(data, indent=2, default=str),
                content_type='application/json'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
            
        elif format == 'csv':
            response = HttpResponse(
                data.getvalue(),
                content_type='text/csv; charset=utf-8'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            
        elif format == 'excel':
            response = HttpResponse(
                data.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            
        elif format == 'pdf':
            response = HttpResponse(
                data.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
            
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        return response
    
    def generate_report(self, report_type: str, start_date=None, end_date=None, 
                       filters: Dict = None) -> Dict:
        """
        Generate analytical reports
        
        Args:
            report_type: Type of report ('summary', 'security', 'performance', 'user_activity')
            start_date: Start date
            end_date: End date
            filters: Additional filters
        
        Returns:
            Report data
        """
        # Set default date range
        if not end_date:
            end_date = timezone.now()
        if not start_date:
            if report_type == 'security':
                start_date = end_date - timedelta(days=7)
            else:
                start_date = end_date - timedelta(days=30)
        
        # Build base queryset
        queryset = AuditLog.objects.filter(
            timestamp__range=(start_date, end_date)
        )
        
        # Apply filters
        if filters:
            from .AuditQuery import AuditQuery
            query_builder = AuditQuery()
            filter_query = query_builder.build_query(filters)
            queryset = queryset.filter(filter_query)
        
        # Generate report based on type
        if report_type == 'summary':
            return self._generate_summary_report(queryset, start_date, end_date)
        elif report_type == 'security':
            return self._generate_security_report(queryset, start_date, end_date)
        elif report_type == 'performance':
            return self._generate_performance_report(queryset, start_date, end_date)
        elif report_type == 'user_activity':
            return self._generate_user_activity_report(queryset, start_date, end_date)
        else:
            raise ValueError(f"Unknown report type: {report_type}")
    
    def _generate_summary_report(self, queryset, start_date, end_date):
        """Generate summary report"""
        
        total = queryset.count()
        
        # Basic statistics
        stats = queryset.aggregate(
            success_count=Count('id', filter=models.Q(success=True)),
            error_count=Count('id', filter=models.Q(level='ERROR')),
            warning_count=Count('id', filter=models.Q(level='WARNING')),
            avg_response_time=Avg('response_time_ms'),
        )
        
        # By action distribution
        actions = queryset.values('action').annotate(
            count=Count('id'),
            avg_time=Avg('response_time_ms'),
            success_rate=Count('id', filter=models.Q(success=True)) * 100.0 / Count('id')
        ).order_by('-count')[:20]
        
        # By user distribution
        users = queryset.filter(user__isnull=False).values(
            'user__email', 'user__username'
        ).annotate(
            count=Count('id'),
            last_activity=Max('timestamp')
        ).order_by('-count')[:10]
        
        # Timeline data
        timeline = queryset.annotate(
            date=TruncDate('timestamp')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        return {
            'report_type': 'summary',
            'period': {
                'start': start_date,
                'end': end_date,
                'days': (end_date - start_date).days
            },
            'summary': {
                'total_actions': total,
                'success_rate': stats['success_count'] / total * 100 if total > 0 else 0,
                'error_count': stats['error_count'],
                'warning_count': stats['warning_count'],
                'avg_response_time_ms': stats['avg_response_time'] or 0,
            },
            'top_actions': list(actions),
            'top_users': list(users),
            'timeline': list(timeline),
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_security_report(self, queryset, start_date, end_date):
        """Generate security report"""
        
        # Filter security-related logs
        security_logs = queryset.filter(
            level__in=['SECURITY', 'ERROR', 'WARNING'],
            action__in=['LOGIN', 'SUSPICIOUS_LOGIN', 'BRUTE_FORCE_ATTEMPT', 
                       'IP_BLOCK', 'SECURITY']
        )
        
        # Failed login attempts
        failed_logins = queryset.filter(
            action='LOGIN',
            success=False
        ).count()
        
        # Suspicious activities
        suspicious = queryset.filter(
            action='SUSPICIOUS_LOGIN'
        ).count()
        
        # IP analysis
        ip_analysis = queryset.values('user_ip').annotate(
            count=Count('id'),
            failed_logins=Count('id', filter=models.Q(action='LOGIN', success=False)),
            suspicious=Count('id', filter=models.Q(action='SUSPICIOUS_LOGIN'))
        ).order_by('-count')[:20]
        
        # User account analysis
        user_analysis = queryset.filter(
            user__isnull=False
        ).values('user__email').annotate(
            total_actions=Count('id'),
            failed_logins=Count('id', filter=models.Q(action='LOGIN', success=False)),
            last_failed_login=Max('timestamp', filter=models.Q(action='LOGIN', success=False))
        ).filter(
            failed_logins__gt=0
        ).order_by('-failed_logins')[:10]
        
        return {
            'report_type': 'security',
            'period': {
                'start': start_date,
                'end': end_date
            },
            'overview': {
                'total_security_events': security_logs.count(),
                'failed_login_attempts': failed_logins,
                'suspicious_activities': suspicious,
                'unique_ips_with_issues': len([ip for ip in ip_analysis if ip['failed_logins'] > 0 or ip['suspicious'] > 0]),
            },
            'ip_analysis': list(ip_analysis),
            'user_analysis': list(user_analysis),
            'recommendations': self._generate_security_recommendations(
                failed_logins, suspicious, ip_analysis
            ),
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_performance_report(self, queryset, start_date, end_date):
        """Generate performance report"""
        
        # Performance metrics
        performance_logs = queryset.exclude(response_time_ms__isnull=True)
        
        stats = performance_logs.aggregate(
            avg_response_time=Avg('response_time_ms'),
            p95_response_time=models.Percentile('response_time_ms', 0.95),
            p99_response_time=models.Percentile('response_time_ms', 0.99),
            max_response_time=Max('response_time_ms'),
            min_response_time=Min('response_time_ms'),
        )
        
        # Slow requests (over 1 second)
        slow_requests = queryset.filter(
            response_time_ms__gt=1000
        ).count()
        
        # By endpoint performance
        endpoint_performance = queryset.exclude(
            request_path__isnull=True
        ).values('request_path', 'request_method').annotate(
            count=Count('id'),
            avg_time=Avg('response_time_ms'),
            max_time=Max('response_time_ms'),
            error_rate=Count('id', filter=models.Q(level='ERROR')) * 100.0 / Count('id')
        ).filter(
            count__gt=10  # Only endpoints with significant traffic
        ).order_by('-avg_time')[:20]
        
        # Performance over time
        hourly_performance = queryset.annotate(
            hour=TruncHour('timestamp')
        ).values('hour').annotate(
            request_count=Count('id'),
            avg_response_time=Avg('response_time_ms'),
            error_count=Count('id', filter=models.Q(level='ERROR'))
        ).order_by('hour')
        
        return {
            'report_type': 'performance',
            'period': {
                'start': start_date,
                'end': end_date
            },
            'overview': {
                'total_requests': queryset.count(),
                'avg_response_time_ms': stats['avg_response_time'] or 0,
                'p95_response_time_ms': stats['p95_response_time'] or 0,
                'p99_response_time_ms': stats['p99_response_time'] or 0,
                'slow_requests': slow_requests,
                'slow_request_percentage': slow_requests / queryset.count() * 100 if queryset.count() > 0 else 0,
            },
            'slow_endpoints': list(endpoint_performance),
            'hourly_performance': list(hourly_performance),
            'recommendations': self._generate_performance_recommendations(
                stats, slow_requests, endpoint_performance
            ),
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_user_activity_report(self, queryset, start_date, end_date):
        """Generate user activity report"""
        
        active_users = queryset.filter(
            user__isnull=False
        ).values('user__id', 'user__email', 'user__username').annotate(
            total_actions=Count('id'),
            first_activity=Min('timestamp'),
            last_activity=Max('timestamp'),
            unique_days=Count('timestamp', distinct=True),
            success_rate=Count('id', filter=models.Q(success=True)) * 100.0 / Count('id')
        ).order_by('-total_actions')
        
        # User engagement over time
        daily_engagement = queryset.annotate(
            date=TruncDate('timestamp')
        ).values('date').annotate(
            unique_users=Count('user', distinct=True),
            total_actions=Count('id')
        ).order_by('date')
        
        # Most active hours
        hourly_activity = queryset.annotate(
            hour=models.ExtractHour('timestamp')
        ).values('hour').annotate(
            action_count=Count('id')
        ).order_by('hour')
        
        # Action patterns
        common_sequences = self._analyze_action_sequences(queryset)
        
        return {
            'report_type': 'user_activity',
            'period': {
                'start': start_date,
                'end': end_date
            },
            'user_summary': {
                'total_active_users': active_users.count(),
                'avg_actions_per_user': active_users.aggregate(avg=Avg('total_actions'))['avg'] or 0,
                'most_active_user': dict(active_users.first()) if active_users.exists() else None,
            },
            'active_users': list(active_users[:50]),
            'daily_engagement': list(daily_engagement),
            'hourly_activity': list(hourly_activity),
            'common_sequences': common_sequences[:10],
            'generated_at': timezone.now().isoformat()
        }
    
    def _analyze_action_sequences(self, queryset):
        """Analyze common action sequences"""
        # This is a simplified version
        # In production, you might use more sophisticated sequence analysis
        
        user_logs = {}
        for log in queryset.filter(user__isnull=False).order_by('timestamp'):
            user_id = log.user_id
            if user_id not in user_logs:
                user_logs[user_id] = []
            user_logs[user_id].append(log.action)
        
        # Find common pairs
        pair_counts = {}
        for actions in user_logs.values():
            for i in range(len(actions) - 1):
                pair = (actions[i], actions[i + 1])
                pair_counts[pair] = pair_counts.get(pair, 0) + 1
        
        # Convert to list
        common_pairs = [
            {'from': pair[0], 'to': pair[1], 'count': count}
            for pair, count in sorted(pair_counts.items(), key=lambda x: x[1], reverse=True)
        ]
        
        return common_pairs
    
    def _generate_security_recommendations(self, failed_logins, suspicious, ip_analysis):
        """Generate security recommendations"""
        recommendations = []
        
        if failed_logins > 100:
            recommendations.append({
                'priority': 'HIGH',
                'title': 'Excessive Failed Login Attempts',
                'description': f'{failed_logins} failed login attempts detected. Consider implementing account lockout or CAPTCHA.',
                'action': 'Review login security settings'
            })
        
        if suspicious > 50:
            recommendations.append({
                'priority': 'HIGH',
                'title': 'Suspicious Activities Detected',
                'description': f'{suspicious} suspicious activities logged. Review security logs and consider additional monitoring.',
                'action': 'Investigate suspicious activities'
            })
        
        # Check for IPs with many failed logins
        risky_ips = [ip for ip in ip_analysis if ip['failed_logins'] > 10]
        if risky_ips:
            recommendations.append({
                'priority': 'MEDIUM',
                'title': 'Risky IP Addresses',
                'description': f'{len(risky_ips)} IP addresses with excessive failed logins detected.',
                'action': 'Consider blocking or rate limiting these IPs',
                'details': [{'ip': ip['user_ip'], 'failed_logins': ip['failed_logins']} for ip in risky_ips[:5]]
            })
        
        if not recommendations:
            recommendations.append({
                'priority': 'LOW',
                'title': 'Security Status Normal',
                'description': 'No critical security issues detected in the audit period.',
                'action': 'Continue regular monitoring'
            })
        
        return recommendations
    
    def _generate_performance_recommendations(self, stats, slow_requests, endpoint_performance):
        """Generate performance recommendations"""
        recommendations = []
        
        avg_response_time = stats.get('avg_response_time', 0)
        if avg_response_time > 500:  # More than 500ms
            recommendations.append({
                'priority': 'HIGH',
                'title': 'High Average Response Time',
                'description': f'Average response time is {avg_response_time:.0f}ms. Consider optimizing database queries or adding caching.',
                'action': 'Review slow database queries and API endpoints'
            })
        
        if slow_requests > 100:
            recommendations.append({
                'priority': 'MEDIUM',
                'title': 'Many Slow Requests',
                'description': f'{slow_requests} requests took more than 1 second to complete.',
                'action': 'Identify and optimize slow endpoints'
            })
        
        # Check for specific slow endpoints
        slow_endpoints = [ep for ep in endpoint_performance if ep['avg_time'] > 1000]
        if slow_endpoints:
            recommendations.append({
                'priority': 'MEDIUM',
                'title': 'Slow Endpoints Detected',
                'description': f'{len(slow_endpoints)} endpoints have average response time > 1 second.',
                'action': 'Optimize these specific endpoints',
                'details': [{'endpoint': f"{ep['request_method']} {ep['request_path']}", 
                           'avg_time': ep['avg_time'], 'calls': ep['count']} 
                          for ep in slow_endpoints[:5]]
            })
        
        if not recommendations:
            recommendations.append({
                'priority': 'LOW',
                'title': 'Performance Status Good',
                'description': 'System performance is within acceptable limits.',
                'action': 'Continue monitoring performance metrics'
            })
        
        return recommendations


# Global instance for convenience
log_exporter = LogExporter()