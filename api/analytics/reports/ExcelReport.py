import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExcelReport:
    """
    Generates Excel reports.
    """

    def __init__(self, title="Report"):
        self.title = title

    def generate_user_analytics_report(self, user_data, output_file="user_analytics_report.xlsx"):
        """
        Generates an Excel report for user analytics.
        user_data: A dictionary containing user analytics data.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Analytics"

        # Title
        ws['A1'] = self.title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        # Date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:B2')

        row = 4

        # Example: Daily Signups Table
        if 'daily_signups' in user_data:
            ws.cell(row=row, column=1, value="Daily Signups")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="Date")
            ws.cell(row=row, column=2, value="Signups")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            row += 1
            for item in user_data['daily_signups']:
                ws.cell(row=row, column=1, value=item['date'])
                ws.cell(row=row, column=2, value=item['count'])
                row += 1
            row += 2

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file)
        return output_file

    def generate_revenue_report(self, revenue_data, output_file="revenue_report.xlsx"):
        """
        Generates an Excel report for revenue analytics.
        revenue_data: A dictionary containing revenue analytics data.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Report"

        # Title
        ws['A1'] = "Revenue Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        # Date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:B2')

        row = 4

        # Example: Daily Revenue Table
        if 'daily_revenue' in revenue_data:
            ws.cell(row=row, column=1, value="Daily Revenue")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value="Date")
            ws.cell(row=row, column=2, value="Revenue")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            row += 1
            for item in revenue_data['daily_revenue']:
                ws.cell(row=row, column=1, value=item['date'])
                ws.cell(row=row, column=2, value=item['total'])
                row += 1
            row += 2

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file)
        return output_file