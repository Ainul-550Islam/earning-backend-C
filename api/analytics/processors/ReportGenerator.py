import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from decimal import Decimal
import pandas as pd
from django.utils import timezone
from django.template.loader import render_to_string
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import json
import os

from ..models import Report
from .DataProcessor import DataProcessor
from ..collectors import (
    UserAnalyticsCollector, RevenueCollector, OfferPerformanceCollector
)

logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    Generate analytics reports in various formats
    """
    
    def __init__(self):
        self.data_processor = DataProcessor()
        self.user_collector = UserAnalyticsCollector()
        self.revenue_collector = RevenueCollector()
        self.offer_collector = OfferPerformanceCollector()
        
        # Templates directory
        self.templates_dir = os.path.join(
            os.path.dirname(__file__),
            '../templates/reports'
        )
    
    def generate_report(
        self,
        report_type: str,
        format: str = 'pdf',
        parameters: Dict = None,
        user = None
    ) -> Report:
        """
        Generate a report
        
        Args:
            report_type: Type of report to generate
            format: Report format (pdf, excel, csv, html, json)
            parameters: Report parameters
            user: User requesting the report
        
        Returns:
            Generated Report instance
        """
        parameters = parameters or {}
        
        try:
            # Collect data based on report type
            report_data = self._collect_report_data(report_type, parameters)
            
            # Generate report file
            file_path, file_size = self._generate_report_file(
                report_type, format, report_data, parameters
            )
            
            # Create Report instance
            report = Report.objects.create(
                name=f"{report_type.replace('_', ' ').title()} - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                report_type=report_type,
                format=format,
                parameters=parameters,
                data=report_data if format == 'json' else {},
                file=file_path,
                file_size=file_size,
                generated_by=user,
                status='completed',
                metadata={
                    'generation_time': timezone.now().isoformat(),
                    'data_points': len(report_data.get('data', [])),
                    'report_type': report_type
                }
            )
            
            logger.info(f"Generated report: {report.name}")
            return report
            
        except Exception as e:
            logger.error(f"Failed to generate report: {str(e)}")
            
            # Create failed report instance
            report = Report.objects.create(
                name=f"{report_type} - Failed",
                report_type=report_type,
                format=format,
                parameters=parameters,
                generated_by=user,
                status='failed',
                metadata={'error': str(e)}
            )
            
            raise
    
    def _collect_report_data(
        self,
        report_type: str,
        parameters: Dict
    ) -> Dict:
        """Collect data for report"""
        
        if report_type == 'daily_summary':
            return self._collect_daily_summary(parameters)
        
        elif report_type == 'weekly_analytics':
            return self._collect_weekly_analytics(parameters)
        
        elif report_type == 'monthly_earnings':
            return self._collect_monthly_earnings(parameters)
        
        elif report_type == 'user_activity':
            return self._collect_user_activity(parameters)
        
        elif report_type == 'revenue_report':
            return self._collect_revenue_report(parameters)
        
        elif report_type == 'offer_performance':
            return self._collect_offer_performance(parameters)
        
        elif report_type == 'referral_report':
            return self._collect_referral_report(parameters)
        
        elif report_type == 'custom':
            return self._collect_custom_report(parameters)
        
        else:
            raise ValueError(f"Unknown report type: {report_type}")
    
    def _collect_daily_summary(self, parameters: Dict) -> Dict:
        """Collect daily summary data"""
        date = parameters.get('date', timezone.now().date())
        
        # Collect various metrics
        user_data = self.user_collector.collect_user_analytics(
            period='daily',
            start_date=datetime.combine(date, datetime.min.time()),
            end_date=datetime.combine(date, datetime.max.time())
        )
        
        revenue_data = self.revenue_collector.calculate_daily_revenue(date)
        
        return {
            'report_type': 'daily_summary',
            'date': date.isoformat(),
            'user_metrics': user_data,
            'revenue_metrics': revenue_data,
            'summary': {
                'total_users': len(set([d['user_id'] for d in user_data if 'user_id' in d])),
                'active_users': revenue_data.get('active_users', 0),
                'total_revenue': revenue_data.get('revenue_total', 0),
                'tasks_completed': sum(d.get('tasks_completed', 0) for d in user_data),
                'offers_completed': sum(d.get('offers_completed', 0) for d in user_data)
            }
        }
    
    def _collect_weekly_analytics(self, parameters: Dict) -> Dict:
        """Collect weekly analytics data"""
        week_start = parameters.get('week_start')
        if not week_start:
            # Default to current week
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
        
        week_end = week_start + timedelta(days=6)
        
        # Collect data
        user_data = self.user_collector.collect_user_analytics(
            period='daily',
            start_date=week_start,
            end_date=week_end
        )
        
        revenue_data = self.revenue_collector.calculate_weekly_revenue(week_start)
        
        # Process trends
        processor = DataProcessor()
        trends = processor.calculate_statistics(
            [{'value': d.get('earnings_total', 0)} for d in user_data],
            'value'
        )
        
        return {
            'report_type': 'weekly_analytics',
            'week_start': week_start.isoformat(),
            'week_end': week_end.isoformat(),
            'user_metrics': user_data,
            'revenue_metrics': revenue_data,
            'trends': trends,
            'insights': self._generate_weekly_insights(user_data, revenue_data)
        }
    
    def _collect_monthly_earnings(self, parameters: Dict) -> Dict:
        """Collect monthly earnings data"""
        month = parameters.get('month')
        if not month:
            month = timezone.now().replace(day=1)
        
        revenue_data = self.revenue_collector.calculate_monthly_revenue(month)
        
        # Add user analytics
        user_data = self.user_collector.collect_user_analytics(
            period='monthly',
            start_date=month,
            end_date=month + timedelta(days=31)
        )
        
        # Calculate per user metrics
        per_user_metrics = {}
        if user_data:
            df = pd.DataFrame(user_data)
            per_user_metrics = {
                'avg_earnings': float(df['earnings_total'].mean()),
                'median_earnings': float(df['earnings_total'].median()),
                'top_earners': df.nlargest(10, 'earnings_total').to_dict('records')
            }
        
        return {
            'report_type': 'monthly_earnings',
            'month': month.strftime('%Y-%m'),
            'revenue_data': revenue_data,
            'user_earnings': user_data,
            'per_user_metrics': per_user_metrics,
            'summary': {
                'total_revenue': revenue_data.get('revenue_total', 0),
                'total_users': len(user_data),
                'avg_earnings_per_user': per_user_metrics.get('avg_earnings', 0),
                'profit_margin': revenue_data.get('profit_margin', 0)
            }
        }
    
    def _collect_user_activity(self, parameters: Dict) -> Dict:
        """Collect user activity data"""
        start_date = parameters.get('start_date', timezone.now() - timedelta(days=30))
        end_date = parameters.get('end_date', timezone.now())
        
        user_data = self.user_collector.collect_user_analytics(
            period='daily',
            start_date=start_date,
            end_date=end_date
        )
        
        # Process for insights
        processor = DataProcessor()
        
        # Segment users
        segments = processor.segment_users(user_data)
        
        # Calculate retention
        retention_data = self.user_collector.collect_user_retention()
        
        return {
            'report_type': 'user_activity',
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'user_data': user_data,
            'segments': segments,
            'retention_data': retention_data,
            'engagement_metrics': self.user_collector.collect_user_engagement()
        }
    
    def _collect_revenue_report(self, parameters: Dict) -> Dict:
        """Collect revenue report data"""
        start_date = parameters.get('start_date', timezone.now() - timedelta(days=90))
        end_date = parameters.get('end_date', timezone.now())
        
        # Collect revenue data
        revenue_by_source = self.revenue_collector.get_revenue_by_source_breakdown(
            start_date, end_date
        )
        
        # Get trends
        trends = self.revenue_collector.calculate_revenue_trends(
            days=(end_date - start_date).days
        )
        
        # Forecast
        processor = DataProcessor()
        forecast_data = processor.forecast_revenue(
            trends.get('daily_data', []),
            periods=30
        )
        
        return {
            'report_type': 'revenue_report',
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'revenue_by_source': revenue_by_source,
            'trends': trends,
            'forecast': forecast_data,
            'key_metrics': {
                'total_revenue': revenue_by_source.get('total_revenue', 0),
                'growth_rate': trends.get('avg_daily_growth', 0),
                'best_day': trends.get('best_performing_day', 'Unknown'),
                'forecast_next_month': forecast_data.get('total_forecast', 0)
            }
        }
    
    def _collect_offer_performance(self, parameters: Dict) -> Dict:
        """Collect offer performance data"""
        offer_id = parameters.get('offer_id')
        start_date = parameters.get('start_date', timezone.now() - timedelta(days=30))
        end_date = parameters.get('end_date', timezone.now())
        
        if offer_id:
            # Single offer report
            performance_data = self.offer_collector.collect_offer_performance(
                offer_id=offer_id,
                start_date=start_date,
                end_date=end_date,
                period='daily'
            )
            
            funnel_data = self.offer_collector.collect_offer_funnel(
                offer_id=offer_id,
                start_date=start_date,
                end_date=end_date
            )
            
            insights = self.offer_collector.collect_offer_insights(
                offer_id=offer_id,
                start_date=start_date,
                end_date=end_date
            )
            
            return {
                'report_type': 'offer_performance',
                'offer_id': offer_id,
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'performance_data': performance_data,
                'funnel_data': funnel_data,
                'insights': insights,
                'is_single_offer': True
            }
        
        else:
            # Multiple offers comparison
            top_offers = self.offer_collector.collect_top_performing_offers(
                metric='revenue',
                limit=20,
                start_date=start_date,
                end_date=end_date
            )
            
            comparison = self.offer_collector.collect_offer_comparison(
                offer_ids=[o['offer_id'] for o in top_offers[:5]],  # Compare top 5
                start_date=start_date,
                end_date=end_date
            )
            
            return {
                'report_type': 'offer_performance',
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'top_offers': top_offers,
                'comparison': comparison,
                'summary_metrics': {
                    'total_offers': len(top_offers),
                    'total_revenue': sum(o['revenue'] for o in top_offers),
                    'avg_conversion_rate': sum(o['conversion_rate'] for o in top_offers) / len(top_offers) if top_offers else 0,
                    'best_offer': max(top_offers, key=lambda x: x['revenue']) if top_offers else None
                },
                'is_single_offer': False
            }
    
    def _collect_referral_report(self, parameters: Dict) -> Dict:
        """Collect referral program data"""
        start_date = parameters.get('start_date', timezone.now() - timedelta(days=90))
        end_date = parameters.get('end_date', timezone.now())
        
        # Get referral events
        from ..models import AnalyticsEvent
        
        referral_events = AnalyticsEvent.objects.filter(
            event_type='referral_joined',
            event_time__gte=start_date,
            event_time__lte=end_date
        )
        
        # Aggregate by referrer
        referral_stats = referral_events.values(
            'metadata__referrer_id',
            'metadata__referrer_username'
        ).annotate(
            referrals_count=Count('id'),
            total_commission=Sum('value'),
            avg_commission=Avg('value')
        ).order_by('-total_commission')
        
        # Calculate program metrics
        total_referrals = referral_events.count()
        total_commission = referral_events.aggregate(total=Sum('value'))['total'] or Decimal('0')
        
        # Get top referrers
        top_referrers = list(referral_stats[:10])
        
        return {
            'report_type': 'referral_report',
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'program_metrics': {
                'total_referrals': total_referrals,
                'total_commission': total_commission,
                'avg_commission_per_referral': total_commission / total_referrals if total_referrals > 0 else 0,
                'active_referrers': len(referral_stats)
            },
            'top_referrers': top_referrers,
            'referral_trends': self._calculate_referral_trends(start_date, end_date)
        }
    
    def _collect_custom_report(self, parameters: Dict) -> Dict:
        """Collect custom report data based on parameters"""
        metrics = parameters.get('metrics', [])
        filters = parameters.get('filters', {})
        
        data = {}
        
        # Collect requested metrics
        for metric in metrics:
            if metric == 'user_activity':
                data['user_activity'] = self._collect_user_activity(parameters)
            elif metric == 'revenue':
                data['revenue'] = self._collect_revenue_report(parameters)
            elif metric == 'offer_performance':
                data['offer_performance'] = self._collect_offer_performance(parameters)
            elif metric == 'referrals':
                data['referrals'] = self._collect_referral_report(parameters)
        
        return {
            'report_type': 'custom',
            'parameters': parameters,
            'data': data,
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_report_file(
        self,
        report_type: str,
        format: str,
        report_data: Dict,
        parameters: Dict
    ) -> Tuple[str, int]:
        """Generate report file in specified format"""
        
        if format == 'pdf':
            return self._generate_pdf_report(report_type, report_data, parameters)
        
        elif format == 'excel':
            return self._generate_excel_report(report_type, report_data, parameters)
        
        elif format == 'csv':
            return self._generate_csv_report(report_type, report_data, parameters)
        
        elif format == 'html':
            return self._generate_html_report(report_type, report_data, parameters)
        
        elif format == 'json':
            return self._generate_json_report(report_type, report_data, parameters)
        
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_pdf_report(
        self,
        report_type: str,
        report_data: Dict,
        parameters: Dict
    ) -> Tuple[str, int]:
        """Generate PDF report"""
        # Render HTML template
        template_name = f"{report_type}.html"
        template_path = os.path.join(self.templates_dir, template_name)
        
        if not os.path.exists(template_path):
            # Use default template
            template_name = "default.html"
        
        context = {
            'report_data': report_data,
            'parameters': parameters,
            'generated_at': timezone.now(),
            'title': report_type.replace('_', ' ').title()
        }
        
        html_content = render_to_string(f"reports/{template_name}", context)
        
        # Generate PDF
        html = HTML(string=html_content)
        pdf_file = html.write_pdf()
        
        # Save file
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join('reports', 'pdf', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        with open(filepath, 'wb') as f:
            f.write(pdf_file)
        
        file_size = os.path.getsize(filepath)
        
        return filepath, file_size
    
    def _generate_excel_report(
        self,
        report_type: str,
        report_data: Dict,
        parameters: Dict
    ) -> Tuple[str, int]:
        """Generate Excel report"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()
        
        # Add header
        header_font = Font(bold=True, size=14)
        ws['A1'] = f"{report_type.replace('_', ' ').title()} Report"
        ws['A1'].font = header_font
        
        # Add metadata
        ws['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Period: {parameters.get('start_date', 'N/A')} to {parameters.get('end_date', 'N/A')}"
        
        # Add data
        row = 6
        
        # Flatten data for Excel
        if 'data' in report_data and isinstance(report_data['data'], list):
            # List of dictionaries
            data_list = report_data['data']
            if data_list:
                # Write headers
                headers = list(data_list[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                
                # Write data
                for item in data_list:
                    row += 1
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        value = item.get(header)
                        cell.value = self._format_excel_value(value)
        
        elif isinstance(report_data, dict):
            # Dictionary data
            for key, value in report_data.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=self._format_excel_value(value))
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('reports', 'excel', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        
        file_size = os.path.getsize(filepath)
        
        return filepath, file_size
    
    def _generate_csv_report(
        self,
        report_type: str,
        report_data: Dict,
        parameters: Dict
    ) -> Tuple[str, int]:
        """Generate CSV report"""
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join('reports', 'csv', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Write CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            if 'data' in report_data and isinstance(report_data['data'], list):
                # List of dictionaries
                data_list = report_data['data']
                if data_list:
                    writer = csv.DictWriter(csvfile, fieldnames=data_list[0].keys())
                    writer.writeheader()
                    writer.writerows(data_list)
            else:
                # Simple dictionary
                writer = csv.writer(csvfile)
                for key, value in report_data.items():
                    writer.writerow([key, value])
        
        file_size = os.path.getsize(filepath)
        
        return filepath, file_size
    
    def _generate_html_report(
        self,
        report_type: str,
        report_data: Dict,
        parameters: Dict
    ) -> Tuple[str, int]:
        """Generate HTML report"""
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        filepath = os.path.join('reports', 'html', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Render HTML
        context = {
            'report_data': report_data,
            'parameters': parameters,
            'generated_at': timezone.now(),
            'title': report_type.replace('_', ' ').title()
        }
        
        html_content = render_to_string('reports/html_template.html', context)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        file_size = os.path.getsize(filepath)
        
        return filepath, file_size
    
    def _generate_json_report(
        self,
        report_type: str,
        report_data: Dict,
        parameters: Dict
    ) -> Tuple[str, int]:
        """Generate JSON report"""
        filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join('reports', 'json', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, default=str)
        
        file_size = os.path.getsize(filepath)
        
        return filepath, file_size
    
    # Helper methods
    def _format_excel_value(self, value):
        """Format value for Excel"""
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, (list, dict)):
            return str(value)
        return value
    
    def _generate_weekly_insights(self, user_data, revenue_data):
        """Generate weekly insights"""
        insights = []
        
        if not user_data:
            return insights
        
        # Calculate metrics
        total_earnings = sum(d.get('earnings_total', 0) for d in user_data)
        avg_engagement = sum(d.get('engagement_score', 0) for d in user_data) / len(user_data)
        
        # Generate insights
        if total_earnings > 10000:
            insights.append("[MONEY] Excellent week! Earnings exceeded $10,000")
        
        if avg_engagement > 70:
            insights.append("[START] High user engagement this week")
        elif avg_engagement < 30:
            insights.append("[WARN] Low user engagement - consider re-engagement campaigns")
        
        if revenue_data.get('profit_margin', 0) > 40:
            insights.append("📈 Strong profit margins this week")
        
        return insights
    
    def _calculate_referral_trends(self, start_date, end_date):
        """Calculate referral trends"""
        from ..models import AnalyticsEvent
        
        # Get daily referral counts
        referrals = AnalyticsEvent.objects.filter(
            event_type='referral_joined',
            event_time__gte=start_date,
            event_time__lte=end_date
        ).annotate(
            date=TruncDate('event_time')
        ).values('date').annotate(
            count=Count('id'),
            total_commission=Sum('value')
        ).order_by('date')
        
        return list(referrals)