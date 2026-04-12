# kyc/reports/generators.py  ── WORLD #1
"""KYC report generators — CSV, Excel, PDF"""
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def generate_kyc_csv(queryset) -> bytes:
    """Generate CSV report from KYC queryset. Returns bytes."""
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header
    writer.writerow([
        'ID', 'Username', 'Email', 'Full Name', 'Status',
        'Document Type', 'Document Number', 'Phone', 'Payment Number',
        'Risk Score', 'Risk Level', 'Is Duplicate', 'Is Face Verified',
        'OCR Confidence', 'Created At', 'Reviewed At', 'Rejection Reason',
    ])

    for kyc in queryset.select_related('user'):
        risk_score = kyc.risk_score or 0
        if risk_score <= 30:   risk_level = 'LOW'
        elif risk_score <= 60: risk_level = 'MEDIUM'
        elif risk_score <= 80: risk_level = 'HIGH'
        else:                  risk_level = 'CRITICAL'

        writer.writerow([
            kyc.id,
            kyc.user.username if kyc.user else '',
            kyc.user.email    if kyc.user else '',
            kyc.full_name,
            kyc.status,
            kyc.document_type,
            kyc.document_number,
            kyc.phone_number,
            kyc.payment_number,
            risk_score,
            risk_level,
            'Yes' if kyc.is_duplicate   else 'No',
            'Yes' if kyc.is_face_verified else 'No',
            f"{kyc.ocr_confidence:.2f}",
            kyc.created_at.strftime('%Y-%m-%d %H:%M') if kyc.created_at else '',
            kyc.reviewed_at.strftime('%Y-%m-%d %H:%M') if kyc.reviewed_at else '',
            kyc.rejection_reason or '',
        ])

    return buf.getvalue().encode('utf-8-sig')   # UTF-8 BOM for Excel compatibility


def generate_kyc_excel(queryset) -> bytes:
    """Generate Excel report. Requires openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed — falling back to CSV")
        return generate_kyc_csv(queryset)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KYC Records'

    # Styles
    header_font    = Font(bold=True, color='FFFFFF', size=11)
    header_fill    = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    center_align   = Alignment(horizontal='center', vertical='center')
    status_colors  = {
        'verified':      'C8E6C9',
        'rejected':      'FFCDD2',
        'pending':       'FFF9C4',
        'not_submitted': 'F5F5F5',
        'expired':       'ECEFF1',
    }

    headers = [
        'ID', 'Username', 'Full Name', 'Status', 'Document Type',
        'Document Number', 'Phone', 'Risk Score', 'Risk Level',
        'Is Duplicate', 'Face Verified', 'Created At',
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell             = ws.cell(row=1, column=col, value=header)
        cell.font        = header_font
        cell.fill        = header_fill
        cell.alignment   = center_align

    # Write data
    for row_idx, kyc in enumerate(queryset.select_related('user'), 2):
        risk_score = kyc.risk_score or 0
        if risk_score <= 30:   risk_level = 'LOW'
        elif risk_score <= 60: risk_level = 'MEDIUM'
        elif risk_score <= 80: risk_level = 'HIGH'
        else:                  risk_level = 'CRITICAL'

        row_data = [
            kyc.id,
            kyc.user.username if kyc.user else '',
            kyc.full_name,
            kyc.status.upper(),
            kyc.document_type,
            kyc.document_number,
            kyc.phone_number,
            risk_score,
            risk_level,
            'YES' if kyc.is_duplicate    else 'NO',
            'YES' if kyc.is_face_verified else 'NO',
            kyc.created_at.strftime('%Y-%m-%d') if kyc.created_at else '',
        ]

        for col, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical='center')

        # Color row by status
        status_color = status_colors.get(kyc.status, 'FFFFFF')
        row_fill     = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = row_fill

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ''))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_kyc_summary_pdf(stats: dict) -> bytes:
    """
    Generate a KYC summary PDF report.
    Requires reportlab. Falls back to plain bytes if not available.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        logger.warning("reportlab not installed — PDF generation skipped")
        return b''

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#1A237E'))
    story.append(Paragraph('KYC Verification Report', title_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 1*cm))

    # Summary table
    table_data = [
        ['Metric',          'Count'],
        ['Total KYC',       stats.get('total', 0)],
        ['Verified',        stats.get('verified', 0)],
        ['Pending',         stats.get('pending', 0)],
        ['Rejected',        stats.get('rejected', 0)],
        ['Expired',         stats.get('expired', 0)],
        ['High Risk',       stats.get('high_risk', 0)],
        ['Duplicates',      stats.get('duplicates', 0)],
        ['Today',           stats.get('submitted_today', 0)],
    ]

    table = Table(table_data, colWidths=[8*cm, 4*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 12),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE',    (0,1), (-1,-1), 10),
        ('TOPPADDING',  (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
    ]))

    story.append(table)
    doc.build(story)
    return buf.getvalue()
